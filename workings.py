#this script reads data from an excel sheet and performs some calculations

#import 'load_workbook' from 'openpyxl' library

from openpyxl import load_workbook

#import pandas and DataFrame
from pandas import DataFrame 
import pandas as pd

#load the workbook containing data
old_workbook = load_workbook(filename="sheet_one.xlsx")
new_sheet = old_workbook.active

#read the file
filename = r"sheet_one.xlsx"
df = pd.read_excel(filename)

#get the user's input values from the columns in the spreadsheet as python lists
name = list(df["NAME"][0:])
gross_pay = list(df["GROSS PAY (N)"][0:])
with_holdings = list(df["TOTAL WITH-HOLDINGS (N)"][0:])
payable_amount = list(df["NET AMOUNT PAYABLE (N)"][0:])


#applying formulae to the values gotten from the spreadsheet
payable_amount = [(gross_pay[i] - with_holdings[i]) for i in range(len(gross_pay)) and range(len(with_holdings))]


#display the results of the above lists as excel columns
df = pd.DataFrame()

#to display final results gotten after performing calculations
df["NAME"] = name[::]
df["GROSS PAY (N)"] = gross_pay[::]
df["TOTAL WITH-HOLDINGS (N)"] = with_holdings[::]
df["NET AMOUNT PAYABLE (N)"] = payable_amount[::]


#then rewrite the values to a new sheet
df.to_excel("EmployeeSalary.xlsx", index=False)




